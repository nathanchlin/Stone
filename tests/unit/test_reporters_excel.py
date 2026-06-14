from datetime import date

import openpyxl

from stone.reporters.excel import ExcelReporter
from stone.selector.engine import SelectionResult
from stone.selector.scoring import StockScore


def _make_result() -> SelectionResult:
    picks = [
        StockScore(
            code="600519",
            name="贵州茅台",
            industry="白酒",
            today=date(2026, 6, 14),
            score=92.3,
            normalized_values={"ma_bullish_alignment": 95.0},
        ),
        StockScore(
            code="000858",
            name="五粮液",
            industry="白酒",
            today=date(2026, 6, 14),
            score=89.7,
            normalized_values={"ma_bullish_alignment": 90.0},
        ),
    ]
    return SelectionResult(
        strategy_name="波段趋势 v1",
        target_date=date(2026, 6, 14),
        universe_size=100,
        scored_size=99,
        passed_size=50,
        final_picks=picks,
    )


def test_excel_reporter_creates_xlsx(tmp_path):
    reporter = ExcelReporter()
    out_path = reporter.render(_make_result(), output_dir=tmp_path)
    assert out_path.exists()
    assert out_path.suffix == ".xlsx"


def test_excel_has_top_picks_sheet(tmp_path):
    reporter = ExcelReporter()
    out_path = reporter.render(_make_result(), output_dir=tmp_path)
    workbook = openpyxl.load_workbook(out_path)
    assert "Top名单" in workbook.sheetnames
    worksheet = workbook["Top名单"]
    assert worksheet.max_row >= 3


def test_excel_has_meta_sheet(tmp_path):
    reporter = ExcelReporter()
    out_path = reporter.render(_make_result(), output_dir=tmp_path)
    workbook = openpyxl.load_workbook(out_path)
    assert "元信息" in workbook.sheetnames


def test_excel_code_column_is_text_format(tmp_path):
    reporter = ExcelReporter()
    out_path = reporter.render(_make_result(), output_dir=tmp_path)
    workbook = openpyxl.load_workbook(out_path)
    worksheet = workbook["Top名单"]
    headers = [cell.value for cell in worksheet[1]]
    code_col_idx = headers.index("代码") + 1
    cell = worksheet.cell(row=2, column=code_col_idx)
    assert cell.data_type == "s" or cell.number_format == "@"


def test_excel_has_disclaimer(tmp_path):
    reporter = ExcelReporter()
    out_path = reporter.render(_make_result(), output_dir=tmp_path)
    workbook = openpyxl.load_workbook(out_path)
    found = False
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                if value and "非投资建议" in str(value):
                    found = True
                    break
    assert found, "Excel report must contain disclaimer"
